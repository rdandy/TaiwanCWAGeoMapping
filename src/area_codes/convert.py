# -*- coding:utf-8 -*-
from __future__ import annotations
import os
from io import BytesIO
import requests
import json
# import pandas as pd
import xlrd
from openpyxl import load_workbook
import xmltodict
import xml
import xml.etree.ElementTree as ET
import argparse

import logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

"""
行政區經緯度：
下載：1050812_行政區經緯度(toPost).xml
download page: https://www.post.gov.tw/post/internet/Download/index.jsp?ID=2292
    查詢：第一層 郵務 > 第二層 郵遞區號(含轉碼軟體) > 8.1 3碼郵遞區號與行政區中心點經緯度對照表

直接連結： https://www.post.gov.tw/post/download/1050812_行政區經緯度%28toPost%29.xml



行政區中英對照：
下載：county_h_10706.xls
download page: https://www.post.gov.tw/post/internet/Download/index.jsp?ID=2292
    查詢：第一層 郵務 > 第二層 郵遞區號(含轉碼軟體) > 6.1 縣市鄉鎮中英對照Excel檔(漢語拼音)

直接連結： https://www.post.gov.tw/post/download/county_h_10706.xls



行政區代碼：
下載：行政區代碼表_Taiwan_Geocode.xlsx
download page: https://alerts.ncdr.nat.gov.tw/CAPfiledownload.aspx
    代碼 > 行政區代碼表

直接連結： https://alerts.ncdr.nat.gov.tw/Document/行政區代碼表_Taiwan_Geocode.xlsx

行政區代碼表_Taiwan_Geocode.xlsx，僅讀取鄉鎮頁籤的 E 欄(Taiwan_Geocode_103_鄉鎮代碼)、G 欄(Taiwan_Geocode_103_縣市鄉鎮名) 欄位


ps: 若發現失效則可能為改版或修正，需要再次查詢以取得新的檔案路徑


輸出 AREA_CODES.json 格式，因有部份地區的郵遞區號是相同的，所以使用從 1 開始為序號的 key 來存放
若 geo_code_103 為空字串，表示該地區在行政區代碼表中找不到，也不會出現在天氣警報系統中
{
    "1": {
        "zip_code": "100",
        "area_name": "臺北市中正區",
        "county_name": "臺北市",
        "city_name": "中正區",
        "longitude": "121.5198722",
        "latitude": "25.03289683",
        "area_name_en": "Zhongzheng Dist., Taipei City",
        "county_name_en": "Taipei City",
        "area_en": "Zhongzheng Dist.",
        "geo_code_103": "6300100"
    },
    "2": {
        "zip_code": "103",
        "area_name": "臺北市大同區",
        "county_name": "臺北市",
        "city_name": "大同區",
        "longitude": "121.5130422",
        "latitude": "25.063424",
        "area_name_en": "Datong Dist., Taipei City",
        "county_name_en": "Taipei City",
        "area_en": "Datong Dist.",
        "geo_code_103": "6300103"
    },
    ...
}
"""
g_path = "https://www.post.gov.tw/post/download/1050812_行政區經緯度(toPost).xml"
xml_file = "1050812_行政區經緯度(toPost).xml"

e_path = "https://www.post.gov.tw/post/download/county_h_10706.xls"
county_en_csv = "county_h_10706.xls"

a_path = "https://alerts.ncdr.nat.gov.tw/Document/行政區代碼表_Taiwan_Geocode.xlsx"
area_geo_code_file = "行政區代碼表_Taiwan_Geocode.xlsx"

area_code_json_file = "AREA_CODES.json"


def fetch_geoxml(file_path: str=g_path) -> list:
    """
    Step 1
    開啟 `行政區經緯度` 檔案，將 xml 轉為 dict 格式
    提取取出檔案中 `dataroot` 內的 `_x0031_050429_行政區經緯度_x0028_toPost_x0029_` 資料
    並依郵遞區號排序

    :param file_path: XML 檔案
    :return: list
    [
        {
            "行政區名": "新竹市北區",
            "_x0033_碼郵遞區號": "300",
            "中心點經度": "120.9491233",
            "中心點緯度": "24.82269542",
            "TGOS_URL": "http://tgos.nat.gov.tw/tgos/Web/MetaData/TGOS_MetaData_View.aspx?MID=9C715A5CD330360D355AE105F908B29E&SHOW_BACK_BUTTON=false"
        },...
    ]
    """

    def _get_ns(_root: xml.etree.ElementTree.Element) -> str:
        ret_ns = ""
        if _root.tag.find("{") == 0 and _root.tag.find("}"):
            ret_ns = xml_root.tag[xml_root.tag.find("{") + 1: xml_root.tag.find("}")]

        return ret_ns

    content = ""
    if file_path.find("http") >= 0:
        logger.debug(f"Download GeoXML from: {file_path}")
        response = requests.get(file_path)
        if response.status_code == 200:
            content = response.content.decode("utf-8")
        else:
            logger.debug(f"Download GeoXML error: {response.status_code}: {response.content}")
            exit(1)
    else:
        if os.path.exists(file_path):
            with open(file_path, "r", encoding="utf-8") as input_file:
                content = input_file.read()
        else:
            logger.debug(f"File not found: {file_path}")
            exit(1)

    area_list = []
    xml_root = None
    if content:
        if content.find("<?xml") != 0:
            logger.error("XML format error")
            exit(1)
        else:
            xml_root = ET.fromstring(content)

        ns = _get_ns(xml_root)
        root_dict = xmltodict.parse(content, process_namespaces=True, namespaces={ns: None})

        area_list = sorted(root_dict.get("dataroot", {}).get("_x0031_050429_行政區經緯度_x0028_toPost_x0029_", []),
                           key=lambda x: x['_x0033_碼郵遞區號'])

    logger.debug(f"Total {len(area_list)} area codes.")
    return area_list


def fetch_enname(file_path: str=e_path) -> list:
    """
    Step 2
    開啟 `county_h_10706.xls` 檔案，將 xls 轉為 list 格式
    提取取出檔案中 `鄉鎮` 頁籤內的 `A` 欄(郵遞區號)、`B` 欄(中文名稱)、`C` 欄(英文名稱) 欄位

    :param file_path: CSV 檔案
    :return: list
    [[100, '臺北市中正區', 'Zhongzheng Dist., Taipei City'], [103, '臺北市大同區', 'Datong Dist., Taipei City'],...]
    """

    if file_path.find("http") >= 0:
        logger.debug(f"Download English Name from: {file_path}")
        response = requests.get(file_path)
        response.raise_for_status()  # 確保請求成功
        if response.status_code == 200:
            # 使用BytesIO讀取下載的內容
            data = BytesIO(response.content)

            # 使用xlrd讀取 XLS 文件
            workbook = xlrd.open_workbook(file_contents=data.read())
            sheet_name = '縣市鄉鎮中英對照檔'
            sheet = workbook.sheet_by_name(sheet_name)

            # 使用pandas讀取 XLS 文件，指定用 usecols 參數只讀取 A、B、C 列
            # df = pd.read_excel(io=data, header=None, sheet_name=0, usecols=[0, 1, 2])
        else:
            logger.error(f"Download English Name error: {response.status_code}: {response.content}")
            exit(1)
    elif os.path.exists(file_path):
        # 使用pandas讀取 XLS 文件，指定用 usecols 參數只讀取 A、B、C 列
        workbook = xlrd.open_workbook(file_path)
        sheet_name = '縣市鄉鎮中英對照檔'
        sheet = workbook.sheet_by_name(sheet_name)

        # df = pd.read_excel(file_path, header=None, sheet_name=0, usecols=[0, 1, 2])
    else:
        logger.error(f"File not found: {file_path}")
        exit(1)

    result_list = []
    # 將 sheet data 轉換 list
    column_indices = [0, 1, 2]
    for row_idx in range(sheet.nrows):
        logger.debug(f"Row {row_idx + 1}: {[sheet.cell_value(row_idx, col_idx) for col_idx in column_indices]}")
        result_list.append([sheet.cell_value(row_idx, col_idx) for col_idx in column_indices])

    # 將 DataFrame 轉換 dlist to list
    # result_list = df.values.tolist()
    return result_list


def fetch_areacode(file_path: str = a_path) -> dict:
    """
    Step 3
    開啟 `行政區代碼表` 檔案，將 excel 轉為 dict 格式
    提取取出檔案中 `鄉鎮` 頁籤內的 `E` 欄(Taiwan_Geocode_103_鄉鎮代碼)、`G` 欄(Taiwan_Geocode_103_縣市鄉鎮名) 欄位

    :param file_path: Excel 檔案
    :return: dict
    {
        "臺北市中正區": {
            "geo_code_103": "6300500",
            "area_name": "中正區",
            "area_name_en": "Jhongjhen District",
            "county_name": "臺北市",
            "county_full_name": "臺北市",
            "county_name_en": "Taipei City",
            "county_geo_code_103": "63"
          },
        "臺北市大同區": {
            "geo_code_103": "6300600",
            "area_name": "大同區",
            "area_name_en": "Datong District",
            "county_name": "臺北市",
            "county_full_name": "臺北市",
            "county_name_en": "Taipei City",
            "county_geo_code_103": "63"
          }
        ...
    }
    """
    logger.debug(f"Fetching area code from: {file_path}")

    if file_path.startswith("http") >= 0:
        logger.debug(f"Download area data from: {file_path}")
        response = requests.get(file_path)
        response.raise_for_status()  # 確保請求成功
        if response.status_code == 200:
            # 使用BytesIO讀取下載的內容
            data = BytesIO(response.content)
            # 打開 xlsx 檔案
            workbook = load_workbook(filename=data)

        else:
            logger.debug(f"Download 行政區代碼表 error: {response.status_code}: {response.content}")
            exit(1)
    elif os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        logger.error(f"File not found: {file_path}")
        exit(1)

    # 使用openpyxl讀取Excel文件
    county_data = {}
    sheet_name = '縣市'
    sheet = workbook[sheet_name]
    # 讀取工作表中的指定欄位 (F, G, H, I)，忽略第一列
    # 縣市代碼|縣市英文名|縣市全名|縣市名
    column_indices = [6, 7, 8, 9]  # F: 6, G: 7, H: 8, I: 9

    # 忽略第一列，從第二列開始讀取資料
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=9, values_only=True):
        logger.debug(row)
        county_data[row[3]] = {
            "geo_code_103": f"{row[0]}",
            "county_name_en": f"{row[1]}",
            "county_full_name": f"{row[2]}"
        }

    result_dict = {}
    sheet_name = '鄉鎮'
    sheet = workbook[sheet_name]
    # 讀取工作表中的指定欄位 (E, F, G, H, I)，忽略第一列
    # 鄉鎮代碼|鄉鎮英文名|縣市鄉鎮名|縣市名|鄉鎮名
    column_indices = [5, 6, 7, 8, 9]  # E: 5, F: 6, G: 7, H: 8, I: 9
    # 忽略第一列，從第二列開始讀取資料
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=9, values_only=True):
        result_dict[row[2]] = {
            'geo_code_103': f"{row[0]}",
            'area_name': row[4],
            'area_name_en': row[1],
            'county_name': row[3],
            'county_full_name': county_data.get(row[3], {}).get("county_full_name", ""),
            'county_name_en': county_data.get(row[3], {}).get("county_name_en", ""),
            'county_geo_code_103': county_data.get(row[3], {}).get("geo_code_103", "")
        }
        logger.debug(f"{row[2]} : {result_dict[row[2]]}")

    return result_dict


def convert(geoxml_path: str = g_path, areacode_path: str = a_path, enname_path: str = e_path,
            out_file: str = area_code_json_file, write_file: bool = True) -> dict:
    """
    Convert the data to AREA_CODES.json format

    :param geoxml_path:
    :param areacode_path:
    :param enname_path:
    :param out_file:
    :param write_file:
    :return: dict
    {
        "1": {
            "zip_code": "100",
            "area_name": "臺北市中正區",
            "area_name_en": "Zhongzheng Dist., Taipei City",
            "county_name": "臺北市",
            "county_geo_code_103": "63",
            "county_full_name": "臺北市",
            "city_name": "中正區",
            "longitude": "121.5198839",
            "latitude": "25.03240487",
            "county_name_en": "Zhongzheng Dist.",
            "area_en": "Taipei City",
            "geo_code_103": "6300500"
        },
        "2": {
            "zip_code": "103",
            "area_name": "臺北市大同區",
            "area_name_en": "Datong Dist., Taipei City",
            "county_name": "臺北市",
            "county_geo_code_103": "63",
            "county_full_name": "臺北市",
            "city_name": "大同區",
            "longitude": "121.5130417",
            "latitude": "25.06342433",
            "county_name_en": "Datong Dist.",
            "area_en": "Taipei City",
            "geo_code_103": "6300600"
        },
        ...
    }
    """
    area_code = {}

    area_list = fetch_geoxml(geoxml_path)
    for c in range(0, len(area_list)):

        area_code[f"{c+1}"] = {
            "zip_code": area_list[c]["_x0033_碼郵遞區號"],
            "area_name": area_list[c]["行政區名"],
            "area_name_en": "",
            "geo_code_103": "",
            "county_name": "",
            "county_name_en": "",
            "county_geo_code_103": "",
            "county_full_name": "",
            "city_name": "",
            "city_name_en": "",
            "longitude": float(area_list[c]["中心點經度"]),
            "latitude":float(area_list[c]["中心點緯度"]),
        }

    enname_list = fetch_enname(enname_path)
    for c in range(0, len(enname_list)):
        for k, v in area_code.items():
            if v.get("area_name") == enname_list[c][1]:
                area_code[k]["area_name_en"] = enname_list[c][2]
                break

    geo_code_103 = fetch_areacode(areacode_path)
    logger.debug(json.dumps(geo_code_103, indent=4, ensure_ascii=False))
    for k, v in area_code.items():
        """
            "geo_code_103": "6300500",
            "area_name": "中正區",
            "area_name_en": "Jhongjhen District",
            "county_name": "臺北市",
            "county_full_name": "臺北市",
            "county_name_en": "Taipei City",
            "county_geo_code_103": "63"
        """
        area_name = v["area_name"]
        area_code[k]["county_name"] = geo_code_103.get(area_name, {}).get("county_name", "")
        area_code[k]["county_name_en"] = geo_code_103.get(area_name, {}).get("county_name_en", "")
        area_code[k]["county_geo_code_103"] = geo_code_103.get(area_name, {}).get("county_geo_code_103", "")
        area_code[k]["county_full_name"] = geo_code_103.get(area_name, {}).get("county_full_name", "")
        area_code[k]["city_name"] = geo_code_103.get(area_name, {}).get("area_name", "")
        area_code[k]["city_name_en"] = geo_code_103.get(area_name, {}).get("area_name_en", "")
        area_code[k]["geo_code_103"] = geo_code_103.get(area_name, {}).get("geo_code_103", "")

    logger.debug(json.dumps(area_code, indent=4, ensure_ascii=False))
    logger.debug(f"Total {len(area_code.keys())} area codes.")

    if write_file:
        try:
            with open(out_file, "w", encoding="utf-8") as output_file:
                json.dump(area_code, output_file, ensure_ascii=False, indent=4)
        except Exception as e:
            logger.error(f"Write file error: {e}")
            exit(1)

    return area_code


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Generate AREA_CODES.json')

    parser.add_argument("-g", "--geoxml", default=g_path,
                        help=f"Specify the GeoXML file path. default: {g_path}")
    parser.add_argument("-a", "--areacode", default=a_path,
                        help=f"Specify the Area Code file path. default: {a_path}")
    parser.add_argument("-e", "--enname", default=e_path,
                        help=f"Specify the English name file path. default: {e_path}")
    parser.add_argument("-o", "--outfile", default=area_code_json_file,
                        help=f"Specify the output file name. default: {area_code_json_file}")

    args = parser.parse_args()

    if args.geoxml:
        g_path = args.geoxml

    if args.areacode:
        a_path = args.areacode

    if args.enname:
        e_path = args.enname

    if args.outfile:
        outfile = args.outfile
    else:
        outfile = area_code_json_file

    result = convert(geoxml_path=g_path,
                     areacode_path=a_path,
                     enname_path=e_path,
                     out_file=outfile,
                     write_file=True)

    logger.debug(json.dumps(result, indent=4, ensure_ascii=False))
